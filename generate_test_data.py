"""
Test Data Generator
====================
INPUT  : schema.xlsx  (all sheets read)

Schema columns (per sheet):
  col 0 : schema_name
  col 1 : table_name
  col 2 : column_name
  col 3 : data_type        (int, varchar, date, decimal, bigint ...)
  col 4 : max_length       (numeric)
  col 5 : precision        (numeric, optional — ignored)
  col 6 : scale            (numeric, optional — ignored)
  col 7 : Rule             (optional — see rule syntax below)

Column positions are auto-detected from the header row (case-insensitive).
The Rule column is found by header name, so precision/scale columns are optional.

Rule syntax (case-insensitive):
  oneof:A,B,C          -> cartesian product dimension, picks each value
  startswith:9033      -> generates value beginning with prefix
  range:1,100          -> boundary values: min, mid, max
  FIXED VALUE          -> always that exact value (plain text, no keyword)
  NULL (in oneof list) -> inserts SQL NULL
  blank                -> smart generation based on datatype / column name

OUTPUT : HMS_Test_Data_<YYYYMMDD_HHMMSS>.xlsx
  Per table:
    <Table>_testdata      - generated rows (blue header)
    <Table>_insertscript  - SQL INSERT statements (green header)
    <Table>_deletescript  - SQL DELETE statements (orange header)

Requires: openpyxl  ->  pip install openpyxl
"""

import itertools
import os
import random
import re
import string
import sys
from collections import OrderedDict
from datetime import date, datetime, timedelta, timezone

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl ...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

try:
    from faker import Faker
except ImportError:
    print("Installing Faker ...")
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'faker'])
    from faker import Faker

fake = Faker()

# ── Excel styles ──────────────────────────────────────────────────────────────

HDR_FONT   = Font(name='Calibri', bold=True, color='FFFFFF')
FILL_BLUE  = PatternFill(fill_type='solid', fgColor='2E75B6')
FILL_GREEN = PatternFill(fill_type='solid', fgColor='375623')
FILL_RED   = PatternFill(fill_type='solid', fgColor='833C00')
HDR_ALIGN  = Alignment(horizontal='center', vertical='center', wrap_text=True)
SQL_FONT   = Font(name='Courier New', size=9)

# ── random helpers ────────────────────────────────────────────────────────────

def rstr(n=8):
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=n))

def ralpha(n=4):
    return ''.join(random.choices(string.ascii_uppercase, k=n))

def rdate():
    s = date(2015, 1, 1)
    return str(s + timedelta(days=random.randint(0, (date(2026, 6, 1) - s).days)))

def rdatetime():
    dt = datetime(2015, 1, 1) + timedelta(days=random.randint(0, 3000),
                                           seconds=random.randint(0, 86399))
    return dt.strftime('%Y-%m-%d %H:%M:%S')

def rdatetimeoffset():
    dt = datetime(2015, 1, 1, tzinfo=timezone.utc) + timedelta(
        days=random.randint(0, 3000), seconds=random.randint(0, 86399))
    return dt.strftime('%Y-%m-%d %H:%M:%S+00:00')

def rdatetime2():
    return rdatetime() + '.0000000'

def rfirst():
    return fake.first_name()

def rlast():
    return fake.last_name()

def rstreet():
    number = random.randint(1, 9999)
    name   = random.choice(['Main','Oak','Maple','Cedar','Pine','Elm','Lake',
                             'Hill','River','Park','Sunset','Willow','Church'])
    suffix = random.choice(['St','Ave','Blvd','Dr','Ln','Rd','Ct','Way'])
    return f"{number} {name} {suffix}"

def rcity():
    return random.choice(['Austin','Houston','Dallas','Phoenix','Denver','Atlanta',
                          'Chicago','Seattle','Boston','Miami','Portland','Orlando',
                          'Nashville','Detroit','Charlotte','Columbus','Memphis'])

def rstate():
    return random.choice(['AL','AK','AZ','AR','CA','CO','CT','DE','FL','GA',
                          'HI','ID','IL','IN','IA','KS','KY','LA','ME','MD',
                          'MA','MI','MN','MS','MO','MT','NE','NV','NH','NJ',
                          'NM','NY','NC','ND','OH','OK','OR','PA','RI','SC',
                          'SD','TN','TX','UT','VT','VA','WA','WV','WI','WY'])

def rzip():
    return f"{random.randint(10000, 99999)}"

def rphone():
    return f"({random.randint(200,999)}) {random.randint(200,999)}-{random.randint(1000,9999)}"

# ── datatype_full builder ─────────────────────────────────────────────────────

def build_datatype_full(data_type: str, max_length, precision, scale) -> str:
    """Reconstruct full datatype string e.g. varchar(30), decimal, int."""
    dt = (data_type or '').lower().strip()
    try: ml = int(max_length)
    except (TypeError, ValueError): ml = 0

    if dt in ('varchar', 'nvarchar', 'char', 'nchar', 'varbinary', 'binary'):
        return f"{dt}({ml})" if ml > 0 else dt
    return dt

def max_len_from_col(col_info: dict) -> int:
    """Return the usable max string length for a column."""
    dt = col_info['datatype_full'].lower()
    m  = re.search(r'\((\d+)', dt)
    if m:
        n = int(m.group(1))
        # nvarchar stores 2 bytes/char; max_length in sys.columns is bytes
        base = re.sub(r'\(.*', '', dt)
        if base in ('nvarchar', 'nchar'):
            return n  # already divided when schema was built via script
        return n
    base = re.sub(r'\(.*', '', dt)
    if base == 'char': return 1
    return 8

# ── schema reader ─────────────────────────────────────────────────────────────

KNOWN_TYPES = {
    'int','bigint','smallint','tinyint','bit',
    'varchar','nvarchar','char','nchar','text','ntext',
    'numeric','decimal','float','real','money','smallmoney',
    'date','datetime','datetime2','datetimeoffset','smalldatetime','time',
    'timestamp','rowversion','uniqueidentifier','xml',
    'varbinary','binary','image'
}

# Header aliases (case-insensitive) for each logical column
_HDR_ALIASES = {
    'schema':     ['schema_name', 'schema', 'schemaname'],
    'table':      ['table_name',  'table',  'tablename'],
    'column':     ['column_name', 'column', 'columnname', 'col_name'],
    'data_type':  ['data_type',   'datatype', 'type'],
    'max_length': ['max_length',  'maxlength', 'length', 'max_len'],
    'rule':       ['rule',        'rules',     'constraint'],
}

def _detect_col_indices(header_row) -> dict:
    """Return {logical_name: index} from the header row."""
    idx_map = {}
    for ci, cell in enumerate(header_row):
        val = str(cell).strip().lower() if cell is not None else ''
        for logical, aliases in _HDR_ALIASES.items():
            if val in aliases and logical not in idx_map:
                idx_map[logical] = ci
    return idx_map


def read_schema(filepath: str) -> dict:
    """
    Reads all sheets.
    Returns { table_name: [ col_info_dict, ... ] }
    col_info_dict keys: column, datatype_full, rule, is_first_col
    Column positions are auto-detected from each sheet's header row.
    """
    wb     = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    tables = {}

    for sheet in wb.worksheets:
        col_idx = {}   # populated from header row

        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            # Detect column positions from header row
            if idx == 0:
                col_idx = _detect_col_indices(row)
                print(f"  [Sheet: {sheet.title}] Column map: {col_idx}")
                continue
            if not row or len(row) < 4:
                continue

            def _get(key, default=None):
                i = col_idx.get(key)
                if i is None or i >= len(row):
                    return default
                return row[i]

            schema_name = str(_get('schema', '') or '').strip()
            table_name  = str(_get('table',  '') or '').strip()
            col_name    = str(_get('column', '') or '').strip()
            data_type   = str(_get('data_type', '') or '').strip()
            max_length  = _get('max_length')
            rule_raw    = _get('rule')
            rule        = str(rule_raw).strip() if rule_raw is not None else ''
            # Treat literal 'None' string (openpyxl empty cell artifact) as blank
            if rule.lower() == 'none':
                rule = ''

            if not table_name or not col_name or not data_type:
                continue

            # Validate datatype
            base_type = data_type.lower().strip()
            if base_type not in KNOWN_TYPES:
                print(f"  [WARN] Skipping [{table_name}].[{col_name}] — "
                      f"unknown type '{data_type}'")
                continue

            datatype_full = build_datatype_full(data_type, max_length, None, None)

            # Track first column per table (used as PK candidate)
            is_first = table_name not in tables
            tables.setdefault(table_name, []).append({
                'column':        col_name,
                'datatype_full': datatype_full,
                'base_type':     base_type,
                'rule':          rule,
                'is_first_col':  is_first,
                'schema':        schema_name,
                'sheet_name':    sheet.title   # ← source Excel sheet name
            })

    wb.close()
    return tables

# ── PK detection ──────────────────────────────────────────────────────────────

def get_pk_col(columns: list):
    """
    Return the PK column name only if the first column is an integer type
    AND does not have a FK rule (FK columns are not PKs of this table).
    """
    if columns:
        first = columns[0]
        if first['base_type'] in ('int', 'bigint', 'smallint', 'tinyint'):
            if not parse_fk_rule(first.get('rule', '')):
                return first['column']
    return None

# ── FK / referential integrity helpers ──────────────────────────────────────

def parse_fk_rule(rule: str):
    """
    Parse FK reference from rule string.
    Supported formats:
      FK -> TableName.ColumnName
      FK -> SchemaName.TableName.ColumnName
    Returns (ref_table, ref_col) or None.
    """
    if not rule:
        return None
    r = rule.strip()
    # 3-part: FK -> schema.table.column
    m = re.match(r'FK\s*->\s*\w+\.(\w+)\.(\w+)', r, re.IGNORECASE)
    if m:
        return (m.group(1), m.group(2))
    # 2-part: FK -> table.column
    m = re.match(r'FK\s*->\s*(\w+)\.(\w+)', r, re.IGNORECASE)
    return (m.group(1), m.group(2)) if m else None


def topo_sort(tables: dict) -> list:
    """Return table names sorted so parent tables come before child tables."""
    # Build case-insensitive name -> actual key map
    name_map = {t.lower(): t for t in tables}
    deps = {t: set() for t in tables}
    for table, columns in tables.items():
        for col in columns:
            fk = parse_fk_rule(col.get('rule', ''))
            if fk:
                actual_parent = name_map.get(fk[0].lower())
                if actual_parent:
                    deps[table].add(actual_parent)

    ordered, visited = [], set()

    def visit(t):
        if t in visited:
            return
        visited.add(t)
        for dep in deps.get(t, []):
            visit(dep)
        ordered.append(t)

    for t in tables:
        visit(t)
    return ordered


# ── rule engine ───────────────────────────────────────────────────────────────

AUTO_GENERATE = object()   # no forced scenario value; use rule/smart generation
FORCE_NULL    = object()   # explicit NULL selected from ONEOF

def rule_scenario_values(rule: str) -> list:
    """
    Return the list of scenario values for cartesian product building.
    - oneof:A,B,C  -> ['A','B','C']  (NULL entry = SQL NULL)
    - range:1,100  -> [1, 50, 100]
    - Anything else (startswith/fixed/blank) -> [AUTO_GENERATE] (single scenario)
    """
    if not rule:
        return [AUTO_GENERATE]
    r = rule.strip()

    m_oneof = re.match(r'^ONEOF\s*:\s*(.*)$', r, re.IGNORECASE)
    if m_oneof:
        raw_vals = [v.strip() for v in m_oneof.group(1).split(',') if v.strip()]
        # Convert 'NULL' token to FORCE_NULL sentinel (-> SQL NULL)
        return [FORCE_NULL if v.upper() == 'NULL' else v for v in raw_vals] or [AUTO_GENERATE]

    m_range = re.match(r'^RANGE\s*:\s*(.*)$', r, re.IGNORECASE)
    if m_range:
        parts = m_range.group(1).split(',')
        if len(parts) == 2:
            try:
                lo, hi = int(parts[0].strip()), int(parts[1].strip())
                return [lo, (lo + hi) // 2, hi]
            except ValueError:
                pass

    return [AUTO_GENERATE]   # startswith / fixed / blank = one scenario


def apply_fixed_rule(rule: str, col_info: dict, forced_value) -> object:
    """
    Given a forced_value from the cartesian product
    (AUTO_GENERATE = generate freely, FORCE_NULL = SQL NULL)
    or a rule string, produce the final cell value.
    """
    ml  = max_len_from_col(col_info)
    dt  = col_info['datatype_full']
    bdt = col_info['base_type']

    # Forced value from ONEOF / RANGE
    if forced_value is FORCE_NULL:
        return None

    if forced_value is not AUTO_GENERATE:
        if isinstance(forced_value, int):
            return forced_value
        return str(forced_value)[:ml]

    # No forced value — apply rule if present
    if not rule:
        return AUTO_GENERATE   # caller uses smart generation

    r  = rule.strip()
    ru = r.upper()

    # FK rules are handled upstream — don't treat as fixed value
    if ru.startswith('FK'):
        return AUTO_GENERATE

    # ONEOF/RANGE are scenario rules, not plain fixed values
    # (this also protects variants like "oneof : NULL")
    if re.match(r'^(ONEOF|RANGE)\s*:', r, re.IGNORECASE):
        return AUTO_GENERATE

    m_startswith = re.match(r'^STARTSWITH\s*:\s*(.*)$', r, re.IGNORECASE)
    if m_startswith:
        prefix     = m_startswith.group(1)
        # Replace each 'x' (case-insensitive) in the prefix with a random digit
        prefix     = re.sub(r'x', lambda _: str(random.randint(0, 9)), prefix, flags=re.IGNORECASE)
        suffix_len = max(0, ml - len(prefix))
        suffix     = ''.join(random.choices(string.ascii_uppercase + string.digits,
                                            k=suffix_len))
        return (prefix + suffix)[:ml]

    m_regex = re.match(r'^REGEX\s*:\s*(.*)$', r, re.IGNORECASE)
    if m_regex:
        return _regex_gen(m_regex.group(1), ml)

    # Plain fixed value
    return str(r)[:ml]


def _regex_gen(pattern: str, max_len: int) -> str:
    result, i = [], 0
    while i < len(pattern):
        ch = pattern[i]
        if ch == '[':
            end   = pattern.index(']', i)
            chars = _expand_cls(pattern[i+1:end])
            i     = end + 1
        elif ch == '\\':
            i += 1
            e = pattern[i] if i < len(pattern) else 'd'
            if   e == 'd': chars = string.digits
            elif e == 'w': chars = string.ascii_letters + string.digits + '_'
            else:          chars = e
            i += 1
        else:
            chars = ch
            i += 1
        qty = 1
        if i < len(pattern):
            if pattern[i] == '{':
                end2 = pattern.index('}', i)
                q    = pattern[i+1:end2]
                if ',' in q:
                    a, b = q.split(',')
                    qty  = random.randint(int(a), int(b))
                else:
                    qty = int(q)
                i = end2 + 1
            elif pattern[i] == '?': qty = random.randint(0, 1); i += 1
            elif pattern[i] == '+': qty = random.randint(1, 5); i += 1
            elif pattern[i] == '*': qty = random.randint(0, 5); i += 1
        result.append(''.join(random.choices(chars, k=qty)))
    return ''.join(result)[:max_len]


def _expand_cls(cls: str) -> str:
    chars, idx = '', 0
    while idx < len(cls):
        if idx + 2 < len(cls) and cls[idx+1] == '-':
            chars += ''.join(chr(c) for c in range(ord(cls[idx]), ord(cls[idx+2])+1))
            idx   += 3
        else:
            chars += cls[idx]; idx += 1
    return chars or string.ascii_uppercase

# ── smart value (no rule) ─────────────────────────────────────────────────────

def smart_value(col_name: str, col_info: dict, pk_counter: list) -> object:
    """Generate a value based on column name + datatype when no rule applies."""
    cl  = col_name.lower()
    bdt = col_info['base_type']
    dt  = col_info['datatype_full']
    ml  = max_len_from_col(col_info)

    # PK auto-increment
    if col_info.get('is_pk'):
        pk_counter[0] += 1
        return pk_counter[0]

    # name hints
    if 'first'  in cl and 'nm' in cl: return rfirst()[:ml]
    if 'last'   in cl and 'nm' in cl: return rlast()[:ml]
    if 'firstname'  in cl: return rfirst()[:ml]
    if 'lastname'   in cl: return rlast()[:ml]
    if 'dob'        in cl or 'birth' in cl: return rdate()
    if 'sex'        in cl: return random.choice(['M','F'])
    if 'street'     in cl or ('addr' in cl and 'line' in cl): return rstreet()[:ml]
    if 'addr'       in cl or 'address' in cl:
        full = rstreet() + ', ' + rcity() + ', ' + rstate()
        # if column is too narrow for full address, fall back to street only
        return full[:ml] if ml >= 30 else rstreet()[:ml]
    if 'city'       in cl: return rcity()[:ml]
    if 'state'      in cl and bdt in ('varchar','char','nvarchar','nchar'): return rstate()[:ml]
    if 'zip'        in cl or 'postal' in cl: return rzip()[:ml]
    if 'phone'      in cl or 'fax' in cl or 'tel' in cl: return rphone()[:ml]
    if 'dos'        == cl or ('date' in cl and 'of' in cl): return rdate()
    if 'status'     in cl and bdt in ('varchar','char','nvarchar'):
        return random.choice(['ACTIVE','PENDING','CLOSED','REVIEW'])[:ml]
    if 'type'       in cl and bdt in ('varchar','char','nvarchar'):
        return ralpha(min(2, ml))
    if 'code'       in cl and bdt in ('varchar','char','nvarchar'):
        return ralpha(min(4, ml))
    if 'flag'       in cl and bdt in ('varchar','char','nvarchar'):
        return random.choice(['Y','N'])
    if 'num'        in cl or 'number' in cl:
        return rstr(min(ml, 10)) if bdt in ('varchar','nvarchar') else random.randint(1, 99999)
    if 'balance'    in cl or 'amount' in cl: return round(random.uniform(0, 9999), 2)
    if 'desc'       in cl: return ('Description ' + rstr(4))[:ml]
    if 'letter'     in cl and bdt in ('varchar','char','nvarchar'): return ralpha(min(4, ml))

    # datatype fallback
    if bdt in ('int','smallint'):                              return random.randint(1, 99999)
    if bdt in ('bigint',):                                     return random.randint(1, 9999999)
    if bdt == 'tinyint':                                       return random.randint(0, 255)
    if bdt == 'bit':                                           return random.randint(0, 1)
    if bdt in ('varchar','nvarchar','char','nchar','text'):    return rstr(min(8, ml))
    if bdt in ('numeric','decimal','float','real','money'):    return round(random.uniform(0,999), 4)
    if bdt == 'date':                                          return rdate()
    if bdt == 'datetime2':                                     return rdatetime2()
    if bdt == 'datetimeoffset':                                return rdatetimeoffset()
    if bdt in ('datetime','smalldatetime'):                    return rdatetime()
    if bdt == 'timestamp':                                     return random.randint(1, 2**32)
    return rstr(min(6, ml))

# ── cartesian product builder ─────────────────────────────────────────────────

def build_combinations(columns: list) -> list:
    """
    Build full cartesian product of all ONEOF/RANGE rule columns.
    Returns list of dicts { col_name: forced_value }.
    AUTO_GENERATE means 'generate using smart/fixed rule'.
    Capped at 500 rows.
    """
    col_names = [c['column'] for c in columns]
    scenarios = [rule_scenario_values(c.get('rule', '')) for c in columns]
    product   = list(itertools.product(*scenarios))

    if len(product) > 500:
        print(f"  [INFO] {len(product)} combinations — capped at 500.")
        product = product[:500]

    return [dict(zip(col_names, combo)) for combo in product]

# ── data generator ────────────────────────────────────────────────────────────

def generate_table_data(table: str, columns: list, pk_start: int,
                        fk_pools: dict = None) -> list:
    pk_col     = get_pk_col(columns)
    pk_counter = [pk_start - 1]   # mutable counter

    # Mark PK column
    for c in columns:
        c['is_pk'] = (c['column'] == pk_col)

    combinations = build_combinations(columns)

    # If this table has FK references, expand rows to cover every parent PK.
    # Find the largest FK pool size among all FK columns in this table.
    max_fk_pool = 0
    if fk_pools:
        for col_info in columns:
            fk = parse_fk_rule(col_info.get('rule', ''))
            if fk:
                ref_table, ref_col = fk
                matched = next((t for t in fk_pools if t.lower() == ref_table.lower()), None)
                if matched:
                    pool = (fk_pools.get(matched) or {}).get(ref_col, [])
                    if not pool:
                        pool = next((v for k, v in fk_pools[matched].items()
                                     if k.lower() == ref_col.lower()), [])
                    if len(pool) > max_fk_pool:
                        max_fk_pool = len(pool)

    # Repeat combinations cyclically until we have enough rows to cover all FK values
    if max_fk_pool > len(combinations):
        base = combinations[:]
        while len(combinations) < max_fk_pool:
            combinations.extend(base)
        combinations = combinations[:max_fk_pool]

    rows = []

    for row_idx, combo in enumerate(combinations):

        row = {}
        for col_info in columns:
            col           = col_info['column']
            rule          = col_info.get('rule', '')
            forced_value  = combo.get(col)   # None = free generation

            if col_info.get('is_pk'):
                pk_counter[0] += 1
                row[col] = pk_counter[0]
                continue

            # Referential integrity: FK column maps row-for-row to parent PK pool
            # row_idx % len(pool) ensures every parent PK is covered in order
            fk = parse_fk_rule(rule)
            if fk and fk_pools:
                ref_table, ref_col = fk
                # case-insensitive table lookup
                matched = next((t for t in fk_pools
                                if t.lower() == ref_table.lower()), None)
                pool = (fk_pools.get(matched) or {}).get(ref_col, []) if matched else []
                if not pool:
                    # try case-insensitive column lookup too
                    if matched:
                        pool = next((v for k, v in fk_pools[matched].items()
                                     if k.lower() == ref_col.lower()), [])
                if pool:
                    row[col] = pool[row_idx % len(pool)]
                    continue
                else:
                    print(f"  [WARN] FK pool empty for {ref_table}.{ref_col} "
                          f"— column [{col}] will get a random value.")

            # Try forced / rule value first
            val = apply_fixed_rule(rule, col_info, forced_value)
            if val is AUTO_GENERATE:
                val = smart_value(col, col_info, pk_counter)
            row[col] = val
        rows.append(row)

    return rows


def generate_all(tables: dict, pk_starts: dict) -> dict:
    all_data = {}
    fk_pools = {}   # { table: { col: [values] } }  — fed to child tables

    generation_order = topo_sort(tables)
    for table in generation_order:
        columns  = tables[table]
        pk_start = pk_starts.get(table, 1)
        print(f"  [{table}] generating ...", end='')
        rows = generate_table_data(table, columns, pk_start, fk_pools)
        print(f" {len(rows)} scenario rows")
        all_data[table] = rows

        # Build FK pool from this table's generated data (all columns)
        fk_pools[table] = {}
        for col_info in columns:
            col  = col_info['column']
            vals = [r[col] for r in rows if r.get(col) is not None]
            if vals:
                fk_pools[table][col] = vals

    # Preserve original table order in output
    return {t: all_data[t] for t in tables if t in all_data}

# ── SQL helpers ───────────────────────────────────────────────────────────────

def sql_val(v) -> str:
    if v is None:            return 'NULL'
    if isinstance(v, bool):  return '1' if v else '0'
    if isinstance(v, int):   return str(v)
    if isinstance(v, float): return str(v)
    return f"'{str(v).replace(chr(39), chr(39)*2)}'"

def build_insert_lines(table: str, schema: str, rows: list, columns: list) -> list:
    pk_col = get_pk_col(columns)
    skip   = {c['column'] for c in columns if c['base_type'] == 'timestamp'}
    if pk_col:
        skip.add(pk_col)   # exclude IDENTITY PK — SQL Server auto-generates it
    cols  = [c['column'] for c in columns if c['column'] not in skip]
    cl    = ', '.join(f'[{c}]' for c in cols)
    lines = [
        f'-- INSERT for [{schema}].[{table}]',
        f'-- Generated : {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        f'-- Rows      : {len(rows)}',
        f'SET NOCOUNT ON;',
        '',
        f'INSERT INTO [{schema}].[{table}] ({cl})',
        f'VALUES'
    ]
    for i, row in enumerate(rows):
        vals   = ', '.join(sql_val(row[c]) for c in cols)
        suffix = ',' if i < len(rows) - 1 else ';'
        lines.append(f'  ({vals}){suffix}')
    lines.append('')
    return lines

def build_delete_lines(table: str, schema: str, rows: list, columns: list) -> list:
    pk_col = get_pk_col(columns)
    lines  = [
        f'-- DELETE for [{schema}].[{table}]',
        f'-- Generated : {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        f'-- Run AFTER child-table deletes if FK exists.',
        f'SET NOCOUNT ON;', ''
    ]
    if pk_col and all(pk_col in r for r in rows):
        pk_vals = ', '.join(sql_val(r[pk_col]) for r in rows)
        lines += [f'DELETE FROM [{schema}].[{table}]',
                  f'WHERE [{pk_col}] IN ({pk_vals});']
    else:
        skip = {c['column'] for c in columns
                if c['base_type'] in ('timestamp', 'bit')}
        for row in rows:
            conds = ' AND '.join(f'[{c}] = {sql_val(v)}'
                                 for c, v in row.items() if c not in skip)
            lines.append(f'DELETE FROM [{schema}].[{table}] WHERE {conds};')
    return lines

# ── Excel writer ──────────────────────────────────────────────────────────────

def sname(table: str, suffix: str) -> str:
    return f"{table}_{suffix}"[:31]

def write_data_sheet(wb, name: str, rows: list):
    ws   = wb.create_sheet(title=name)
    cols = list(rows[0].keys())
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(1, ci, c)
        cell.font, cell.fill, cell.alignment = HDR_FONT, FILL_BLUE, HDR_ALIGN
    for ri, row in enumerate(rows, 2):
        for ci, c in enumerate(cols, 1):
            ws.cell(ri, ci, row[c])
    for ci, c in enumerate(cols, 1):
        w = max(len(c), max((len(str(r[c])) for r in rows if r[c] is not None), default=4))
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 2, 40)
    ws.freeze_panes = 'A2'

def write_sql_sheet(wb, name: str, lines: list, fill):
    ws = wb.create_sheet(title=name)
    cell = ws.cell(1, 1, name)
    cell.font, cell.fill, cell.alignment = HDR_FONT, fill, HDR_ALIGN
    for ri, line in enumerate(lines, 2):
        cell = ws.cell(ri, 1, line)
        cell.font = SQL_FONT
    ws.column_dimensions['A'].width = 120

def write_workbook(all_data: dict, tables: dict, output_dir: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Group tables by their source schema sheet name
    sheet_groups = OrderedDict()
    for table, rows in all_data.items():
        if not rows:
            continue
        src_sheet = tables[table][0].get('sheet_name', 'Sheet1')
        sheet_groups.setdefault(src_sheet, []).append(table)

    for src_sheet, table_list in sheet_groups.items():
        # Sheet name prefix capped so full name fits Excel's 31-char limit
        prefix    = src_sheet[:12]
        s_data    = f"{prefix}_Test Data Preview"[:31]
        s_insert  = f"{prefix}_Insert Scripts"[:31]
        s_delete  = f"{prefix}_Delete Scripts"[:31]

        # ── Data Preview sheet ────────────────────────────────────────────────
        ws_data     = wb.create_sheet(title=s_data)
        current_row = 1

        for table in table_list:
            rows = all_data[table]

            # Table name banner
            banner           = ws_data.cell(current_row, 1, f'[ {table} ]  —  {len(rows)} row(s)')
            banner.font      = Font(name='Calibri', bold=True, color='FFFFFF')
            banner.fill      = PatternFill(fill_type='solid', fgColor='1F4E79')
            banner.alignment = Alignment(horizontal='left', vertical='center')
            current_row     += 1

            # Header row
            cols = list(rows[0].keys())
            for ci, c in enumerate(cols, 1):
                cell = ws_data.cell(current_row, ci, c)
                cell.font, cell.fill, cell.alignment = HDR_FONT, FILL_BLUE, HDR_ALIGN
            current_row += 1

            # Data rows
            for row in rows:
                for ci, c in enumerate(cols, 1):
                    ws_data.cell(current_row, ci, row[c])
                current_row += 1

            current_row += 1   # blank separator

        # Auto-fit columns
        for col_cells in ws_data.columns:
            max_w = max((len(str(c.value)) for c in col_cells if c.value), default=8)
            ws_data.column_dimensions[
                get_column_letter(col_cells[0].column)].width = min(max_w + 2, 40)
        ws_data.freeze_panes = 'A2'

        # ── Insert Scripts sheet ──────────────────────────────────────────────
        ws_ins  = wb.create_sheet(title=s_insert)
        ins_row = 1
        hdr     = ws_ins.cell(ins_row, 1, s_insert)
        hdr.font, hdr.fill, hdr.alignment = HDR_FONT, FILL_GREEN, HDR_ALIGN
        ins_row += 1

        for table in table_list:
            rows    = all_data[table]
            columns = tables[table]
            schema  = columns[0].get('schema', 'dbo')
            for line in build_insert_lines(table, schema, rows, columns):
                cell      = ws_ins.cell(ins_row, 1, line)
                cell.font = SQL_FONT
                ins_row  += 1
            ins_row += 1

        ws_ins.column_dimensions['A'].width = 120

        # ── Delete Scripts sheet ──────────────────────────────────────────────
        ws_del  = wb.create_sheet(title=s_delete)
        del_row = 1
        hdr2    = ws_del.cell(del_row, 1, s_delete)
        hdr2.font, hdr2.fill, hdr2.alignment = HDR_FONT, FILL_RED, HDR_ALIGN
        del_row += 1

        for table in table_list:
            rows    = all_data[table]
            columns = tables[table]
            schema  = columns[0].get('schema', 'dbo')
            for line in build_delete_lines(table, schema, rows, columns):
                cell      = ws_del.cell(del_row, 1, line)
                cell.font = SQL_FONT
                del_row  += 1
            del_row += 1

        ws_del.column_dimensions['A'].width = 120

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out = os.path.join(output_dir, f'HMS_Test_Data_{timestamp}.xlsx')
    wb.save(out)
    print(f"\n  -> Workbook saved : {out}")
    print(f"     Sheets         : {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"       {s}")

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    schema_file = os.path.join(script_dir, 'HMS_Table_Schema.xlsx')

    if not os.path.exists(schema_file):
        print("ERROR: Schema file 'HMS_Table_Schema.xlsx' not found in:")
        print(f"  {script_dir}")
        print("Please place the file in the same folder as this script and re-run.")
        return

    print("=" * 60)
    print("  Test Data Generator")
    print("=" * 60)
    print(f"Schema : {schema_file}\n")

    tables = read_schema(schema_file)
    if not tables:
        print("No tables found in schema.")
        return

    print(f"Tables found : {', '.join(tables.keys())}\n")
    print("Record count per table is auto-calculated from ONEOF/RANGE rules")
    print("(full cartesian product, capped at 500 rows per table).\n")

    # Per-table PK start prompt (only if table has columns)
    print("Enter the CURRENT MAX existing PK for each table.")
    print("  Enter 0 if the table is empty.\n")

    pk_starts = {}
    for table, columns in tables.items():
        pk_col = get_pk_col(columns)
        if not pk_col:
            print(f"  [{table}] — no PK detected, skipping.\n")
            continue
        print(f"  Run in DB : SELECT MAX([{pk_col}]) FROM [{columns[0]['schema']}].[{table}];")
        while True:
            try:
                val = int(input(f"  Current max PK in [{table}] (0 if empty): ").strip())
                if val >= 0:
                    pk_starts[table] = val + 1
                    print(f"    -> PKs will start at {pk_starts[table]}\n")
                    break
                print("  Enter 0 or a positive integer.")
            except ValueError:
                print("  Invalid input.")

    print("Generating scenario combinations ...\n")
    all_data = generate_all(tables, pk_starts)
    write_workbook(all_data, tables, script_dir)
    print("\nDone!")

if __name__ == '__main__':
    main()
